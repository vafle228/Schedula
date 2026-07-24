"""Curriculum (teaching-load) Excel export.

Fills the state-issued «Расчёт часов годовой учебной нагрузки» template with plan
data: one row per *(teacher, discipline)*, grouped by teacher with a per-teacher
«Итого:» subtotal, a department-wide «Итого за кафедру:» total and the template's
footer notes / signatures. The template header (rows 1-18 — the frozen table head)
is preserved verbatim; only the body from row 19 down is regenerated.

Hours land in the semester block (autumn G-S / spring U-AG) dictated by the
discipline's ``period``; the topic-type key selects the column within the block.
"""

from __future__ import annotations

import io
import re
from collections.abc import Iterable, Iterator, Mapping
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models.assignment import Assignment
from core.models.discipline import Discipline
from core.models.group import Group
from core.models.major import Major
from core.models.teacher import Teacher

_TEMPLATE_PATH = Path(__file__).parent / "templates" / "curriculum_template.xlsx"
_SHEET_NAME = "Лист1"

# First column of each semester block; the two blocks are column-aligned so a
# fall column maps to its spring twin by a fixed offset.
_FALL_START = column_index_from_string("G")
_SPRING_START = column_index_from_string("U")

# Fixed non-data columns (1-based).
_COL_INDEX = 1  # A — №
_COL_TEACHER = 2  # B — должность, фамилия, инициалы
_COL_DISCIPLINE = 3  # C — наименование дисциплин
_COL_YEAR = 4  # D — год обучения
_COL_SPECIALTY = 5  # E — специальность
_COL_GROUPS = 6  # F — количество учебных групп
_COL_COURSEWORK = column_index_from_string("AI")  # Рук КЗ, КР (не по семестрам)
_LAST_COLUMN = column_index_from_string("AM")  # всего за учебный год

# Topic-type key -> fall column letter; the spring twin is derived by offset.
# «со группой / с п/группой» splits collapse onto the first (group) column.
_KIND_FALL_COLUMN: dict[str, str] = {
    "lec": "G",  # Лекции
    "grp": "H",  # Групповые занятия
    "prac": "I",  # Практические занятия (со группой)
    "lab": "J",  # Практические занятия (с п/группой)
    "grpex": "K",  # Групповые упражнения (со группой)
    "sem": "M",  # Семинары
    "control": "N",  # Контрольные работы
    "srs": "O",  # Руководство самостоятельной работой студентов
    "train": "P",  # Тренировки
    "consult": "Q",  # Консультации
    "credit": "R",  # Зачёты
    "exam": "S",  # Экзамены
}
# Kinds booked to the single, non-semester course-work column (Рук КЗ, КР).
_COURSEWORK_KINDS = frozenset({"course"})

_FIRST_BODY_ROW = 19
# Rows in the pristine template we clone per-column cell styling from.
_PROTO_DATA_ROW = 19
_PROTO_SUBTOTAL_ROW = 23
_PROTO_TOTAL_ROW = 24

# Columns carrying a vertically-summable value (skip the computed ИТОГО/ВСЕГО).
_VERTICAL_SUM_COLUMNS: tuple[int, ...] = (
    *range(column_index_from_string("G"), column_index_from_string("S") + 1),
    *range(column_index_from_string("U"), column_index_from_string("AG") + 1),
    column_index_from_string("AI"),
    column_index_from_string("AK"),
    column_index_from_string("AL"),
)
# Every numeric column of the department-total row.
_TOTAL_COLUMNS: tuple[int, ...] = tuple(
    range(column_index_from_string("G"), _LAST_COLUMN + 1)
)

_DATA_ROW_HEIGHT = 15.75
_SUBTOTAL_ROW_HEIGHT = 23.25
_TOTAL_ROW_HEIGHT = 15.0


def _spring_column(fall_letter: str) -> str:
    """Return the spring-block column twin of a fall-block column letter."""
    idx = column_index_from_string(fall_letter) - _FALL_START + _SPRING_START
    return get_column_letter(idx)


def _column_for(kind: str, period: str) -> str | None:
    """Return the target column letter for ``kind`` in a ``period`` discipline.

    Args:
        kind: Topic-type key (``lec`` / ``prac`` / …).
        period: Discipline season, ``"fall"`` or ``"spring"``.

    Returns:
        The column letter, or ``None`` for a kind with no template column.
    """
    if kind in _COURSEWORK_KINDS:
        return get_column_letter(_COL_COURSEWORK)
    fall = _KIND_FALL_COLUMN.get(kind)
    if fall is None:
        return None
    return fall if period == "fall" else _spring_column(fall)


@dataclass(slots=True)
class _PlanRow:
    """One rendered body row: a discipline within a single teacher's load."""

    teacher_id: int
    teacher_name: str
    discipline_name: str
    year: int
    specialty: str
    group_count: int
    hours: dict[str, int] = field(default_factory=dict)


@dataclass(slots=True)
class _Bucket:
    """Mutable accumulator for a ``(teacher, discipline)`` pair."""

    discipline: Discipline
    groups: set[int] = field(default_factory=set)
    hours: dict[str, int] = field(default_factory=dict)


def _collect_rows(
    period: str,
    teachers: list[Teacher],
    disciplines: list[Discipline],
    groups: list[Group],
    assignments: Mapping[tuple[int, int], Assignment],
    majors: list[Major],
) -> list[_PlanRow]:
    """Aggregate assignments into ``(teacher, discipline)`` body rows.

    Hours are summed over every ``(group, topic)`` the teacher is assigned, so a
    topic taught to several groups contributes its hours once per group.
    """
    topic_index: dict[int, Discipline] = {}
    topic_hours: dict[int, tuple[str, int]] = {}
    for discipline in disciplines:
        for topic in discipline.topics:
            topic_index[topic.id] = discipline
            topic_hours[topic.id] = (topic.kind, topic.hours)

    major_code = {major.id: major.code for major in majors}
    teacher_order = {teacher.id: idx for idx, teacher in enumerate(teachers)}
    teacher_name = {teacher.id: teacher.name for teacher in teachers}

    buckets: dict[tuple[int, int], _Bucket] = {}
    for (group_id, topic_id), assignment in assignments.items():
        discipline = topic_index.get(topic_id)
        if discipline is None:
            continue
        if period != "both" and discipline.period != period:
            continue
        kind, hours = topic_hours[topic_id]
        column = _column_for(kind, discipline.period)
        if column is None:
            continue
        key = (assignment.teacher_id, discipline.id)
        bucket = buckets.get(key)
        if bucket is None:
            bucket = buckets[key] = _Bucket(discipline=discipline)
        bucket.groups.add(group_id)
        bucket.hours[column] = bucket.hours.get(column, 0) + hours

    rows = [
        _PlanRow(
            teacher_id=teacher_id,
            teacher_name=teacher_name.get(teacher_id, "—"),
            discipline_name=bucket.discipline.name,
            year=bucket.discipline.course,
            specialty=major_code.get(bucket.discipline.major_id, ""),
            group_count=len(bucket.groups),
            hours=bucket.hours,
        )
        for (teacher_id, _discipline_id), bucket in buckets.items()
    ]
    rows.sort(key=lambda row: (teacher_order.get(row.teacher_id, 1 << 30), row.discipline_name))
    return rows


def _group_by_teacher(rows: list[_PlanRow]) -> Iterator[list[_PlanRow]]:
    """Yield consecutive runs of rows that share a teacher (rows are presorted)."""
    block: list[_PlanRow] = []
    for row in rows:
        if block and row.teacher_id != block[0].teacher_id:
            yield block
            block = []
        block.append(row)
    if block:
        yield block


def _capture_row_styles(ws: Worksheet, row: int) -> dict[int, object]:
    """Clone the per-column cell style of a template ``row`` for later reuse."""
    return {
        col: copy(ws.cell(row=row, column=col)._style)
        for col in range(1, _LAST_COLUMN + 1)
    }


def _apply_row_styles(ws: Worksheet, row: int, styles: Mapping[int, object]) -> None:
    """Stamp captured per-column styles onto a freshly written ``row``."""
    for col, style in styles.items():
        ws.cell(row=row, column=col)._style = copy(style)


def _sum_formula(letter: str, first: int, last: int) -> str:
    """Blank-aware vertical sum of a column between two body rows."""
    span = f"{letter}{first}:{letter}{last}"
    return f'=IF(SUM({span})=0,"",SUM({span}))'


def _row_total_formulas(ws: Worksheet, row: int) -> None:
    """Write the per-row ИТОГО / ВСЕГО roll-up formulas shared by every row."""
    ws[f"T{row}"] = f'=IF(SUM(G{row}:S{row})=0,"",SUM(G{row}:S{row}))'
    ws[f"AH{row}"] = f'=IF(SUM(U{row}:AG{row})=0,"",SUM(U{row}:AG{row}))'
    ws[f"AJ{row}"] = f'=IF(SUM(T{row},AH{row})=0,"",SUM(T{row},AH{row}))'
    ws[f"AM{row}"] = f'=IF(SUM(AJ{row}:AL{row})=0,"",SUM(AJ{row}:AL{row}))'


def _write_data_row(
    ws: Worksheet,
    row: int,
    styles: Mapping[int, object],
    plan: _PlanRow,
    *,
    ordinal: int | None,
    teacher: str | None,
) -> None:
    """Render one discipline row; ``ordinal``/``teacher`` set only on a block's head."""
    _apply_row_styles(ws, row, styles)
    ws.row_dimensions[row].height = _DATA_ROW_HEIGHT
    if ordinal is not None:
        ws.cell(row=row, column=_COL_INDEX).value = ordinal
    if teacher is not None:
        ws.cell(row=row, column=_COL_TEACHER).value = teacher
    ws.cell(row=row, column=_COL_DISCIPLINE).value = plan.discipline_name
    ws.cell(row=row, column=_COL_YEAR).value = plan.year
    ws.cell(row=row, column=_COL_SPECIALTY).value = plan.specialty
    ws.cell(row=row, column=_COL_GROUPS).value = plan.group_count
    for letter, hours in plan.hours.items():
        ws[f"{letter}{row}"].value = hours
    _row_total_formulas(ws, row)


def _write_subtotal_row(
    ws: Worksheet,
    row: int,
    styles: Mapping[int, object],
    first: int,
    last: int,
) -> None:
    """Render a per-teacher «Итого:» subtotal over body rows ``first..last``."""
    _apply_row_styles(ws, row, styles)
    ws.row_dimensions[row].height = _SUBTOTAL_ROW_HEIGHT
    ws.cell(row=row, column=_COL_TEACHER).value = "Итого:"
    for col in _VERTICAL_SUM_COLUMNS:
        letter = get_column_letter(col)
        ws[f"{letter}{row}"] = _sum_formula(letter, first, last)
    _row_total_formulas(ws, row)


def _write_total_row(
    ws: Worksheet,
    row: int,
    styles: Mapping[int, object],
    subtotal_rows: Iterable[int],
) -> None:
    """Render the department «Итого за кафедру:» total over the subtotal rows."""
    _apply_row_styles(ws, row, styles)
    ws.row_dimensions[row].height = _TOTAL_ROW_HEIGHT
    ws.cell(row=row, column=_COL_TEACHER).value = "Итого за кафедру:"
    for col in _TOTAL_COLUMNS:
        letter = get_column_letter(col)
        refs = ",".join(f"{letter}{sr}" for sr in subtotal_rows) or "0"
        ws[f"{letter}{row}"] = f'=IF(SUM({refs})=0," ",SUM({refs}))'


def _update_title(ws: Worksheet, year_label: str) -> None:
    """Refresh the academic-year span inside the report title (cell A11)."""
    title = ws["A11"].value
    if isinstance(title, str):
        ws["A11"].value = re.sub(r"\d{4}\s*[-–/]\s*\d{2,4}", year_label, title, count=1)


def _clear_body(ws: Worksheet) -> None:
    """Drop every row (and merged range) from the first body row downwards."""
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= _FIRST_BODY_ROW:
            ws.unmerge_cells(str(merged))
    if ws.max_row >= _FIRST_BODY_ROW:
        ws.delete_rows(_FIRST_BODY_ROW, ws.max_row - _FIRST_BODY_ROW + 1)


def _write_footer(ws: Worksheet, row: int, footer: "_Footer") -> None:
    """Append the note line and signature block below the totals."""
    note_row = row + 1
    ws.cell(row=note_row, column=_COL_TEACHER).value = footer.note_text
    ws.cell(row=note_row, column=_COL_TEACHER)._style = copy(footer.note_style)
    ws.row_dimensions[note_row].height = 16.5
    ws.merge_cells(
        start_row=note_row, start_column=_COL_TEACHER, end_row=note_row, end_column=_LAST_COLUMN
    )

    head_row = note_row + 2
    ws.cell(row=head_row, column=_COL_TEACHER).value = footer.head_text
    ws.cell(row=head_row, column=_COL_TEACHER)._style = copy(footer.head_style)
    ws.row_dimensions[head_row].height = 22.05

    sign_row = head_row + 1
    ws.cell(row=sign_row, column=_COL_TEACHER).value = footer.sign_text
    ws.cell(row=sign_row, column=_COL_TEACHER)._style = copy(footer.sign_style)
    val_col = column_index_from_string("AB")
    ws.cell(row=sign_row, column=val_col).value = footer.sign_value_text
    ws.cell(row=sign_row, column=val_col)._style = copy(footer.sign_value_style)
    ws.row_dimensions[sign_row].height = 22.05
    ws.merge_cells(
        start_row=sign_row, start_column=_COL_TEACHER, end_row=sign_row, end_column=8
    )
    ws.merge_cells(
        start_row=sign_row, start_column=val_col, end_row=sign_row, end_column=_LAST_COLUMN
    )


@dataclass(slots=True)
class _Footer:
    """Footer text and styling captured from the template before the wipe."""

    note_text: object
    note_style: object
    head_text: object
    head_style: object
    sign_text: object
    sign_style: object
    sign_value_text: object
    sign_value_style: object


def _capture_footer(ws: Worksheet) -> _Footer:
    """Snapshot the note/signature cells so the wipe can regenerate them."""
    return _Footer(
        note_text=ws["B26"].value,
        note_style=copy(ws["B26"]._style),
        head_text=ws["B28"].value,
        head_style=copy(ws["B28"]._style),
        sign_text=ws["B29"].value,
        sign_style=copy(ws["B29"]._style),
        sign_value_text=ws["AB29"].value,
        sign_value_style=copy(ws["AB29"]._style),
    )


def build_curriculum_workbook(
    *,
    year_label: str,
    period: str,
    teachers: list[Teacher],
    disciplines: list[Discipline],
    groups: list[Group],
    assignments: Mapping[tuple[int, int], Assignment],
    majors: list[Major],
) -> bytes:
    """Build the filled teaching-load workbook and return it as ``.xlsx`` bytes.

    Args:
        year_label: Academic-year span for the title, e.g. ``"2026-2027"``.
        period: ``"fall"``, ``"spring"`` or ``"both"``.
        teachers: All teachers (also fixes the block order in the sheet).
        disciplines: Disciplines with their embedded topics.
        groups: Groups (unused beyond assignment counting, kept for symmetry).
        assignments: ``(group_id, topic_id) -> Assignment`` map for the year.
        majors: Majors, for the specialty (E) column.

    Returns:
        The serialized ``.xlsx`` document.
    """
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[_SHEET_NAME]
    for name in list(wb.sheetnames):
        if name != _SHEET_NAME:
            del wb[name]

    proto_data = _capture_row_styles(ws, _PROTO_DATA_ROW)
    proto_subtotal = _capture_row_styles(ws, _PROTO_SUBTOTAL_ROW)
    proto_total = _capture_row_styles(ws, _PROTO_TOTAL_ROW)
    footer = _capture_footer(ws)

    _update_title(ws, year_label)
    _clear_body(ws)

    rows = _collect_rows(period, teachers, disciplines, groups, assignments, majors)

    row = _FIRST_BODY_ROW
    subtotal_rows: list[int] = []
    for ordinal, block in enumerate(_group_by_teacher(rows), start=1):
        first = row
        for offset, plan in enumerate(block):
            _write_data_row(
                ws,
                row,
                proto_data,
                plan,
                ordinal=ordinal if offset == 0 else None,
                teacher=plan.teacher_name if offset == 0 else None,
            )
            row += 1
        _write_subtotal_row(ws, row, proto_subtotal, first, row - 1)
        subtotal_rows.append(row)
        row += 1

    _write_total_row(ws, row, proto_total, subtotal_rows)
    _write_footer(ws, row, footer)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
