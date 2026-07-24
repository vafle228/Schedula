"""Weekly-schedule Excel export.

Renders the placed lessons of one semester into the institution's «РАСПИСАНИЕ
ЗАНЯТИЙ» form. The logical structure mirrors the reference template verbatim:

* one **sheet per teaching day** (ПОНЕДЕЛЬНИК … ВОСКРЕСЕНЬЕ);
* inside a day, one **block per entity** — a group (group view) or a teacher
  (teacher view) — stacked vertically;
* the block's left columns carry the bell grid (``Учебный час`` / ``время``),
  one **row-group of four lines per bell slot** — *занятие / тема + тип /
  кабинет / ФИО* — exactly as the paper form;
* every **teaching week is a column**, headed by that day's calendar date.

The four stacked lines of a filled cell read, top to bottom: the discipline, the
topic label with its lesson-type short code, the room and — depending on the
view — the teacher (group view) or the group (teacher view).
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models.discipline import Discipline
from core.models.group import Group
from core.models.lesson import Lesson
from core.models.major import Major
from core.models.settings import SemesterSettings
from core.models.teacher import Teacher
from core.models.topic_type import TopicType

# Monday-first RU day names; the sheet (and its header) is titled by this.
_DAY_NAMES: tuple[str, ...] = (
    "ПОНЕДЕЛЬНИК",
    "ВТОРНИК",
    "СРЕДА",
    "ЧЕТВЕРГ",
    "ПЯТНИЦА",
    "СУББОТА",
    "ВОСКРЕСЕНЬЕ",
)
_SEASON_LABEL: dict[str, str] = {"fall": "осеннем", "spring": "весеннем"}

# Fixed left columns; the week columns start at :data:`_FIRST_WEEK_COL`.
_COL_GROUP = 1  # A — Учебная группа / block label
_COL_HOUR = 2  # B — Учебный час (pair number)
_COL_TIME = 3  # C — время
_FIRST_WEEK_COL = 4  # D … — one per teaching week

# One bell slot occupies four stacked lines in a block.
_LINES_PER_SLOT = 4

_DATA_FILL = PatternFill("solid", fgColor="FFFFE4E1")

_MED = Side(style="medium", color="FF000000")
_THIN = Side(style="thin", color="FF000000")
_NONE = Side()

_FONT_TITLE = Font(name="Times New Roman", size=20)
_FONT_DAY = Font(name="Times New Roman", size=24, bold=True)
_FONT_HEAD = Font(name="Times New Roman", size=12, bold=True)
_FONT_LABEL = Font(name="Times New Roman", size=13, bold=True)
_FONT_HOUR = Font(name="Times New Roman", size=12, bold=True)
_FONT_DISC = Font(name="Times New Roman", size=11, bold=True)
_FONT_CELL = Font(name="Times New Roman", size=10)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass(slots=True)
class _Entity:
    """One rendered block: a group or a teacher, plus its column A caption."""

    key: int  # group id (group view) or teacher id (teacher view)
    caption: str


@dataclass(slots=True)
class _CellText:
    """The four stacked lines of a single (slot, week) schedule cell."""

    discipline: str
    topic: str
    room: str
    person: str

    def line(self, index: int) -> str:
        """Return the ``index``-th (0-based, top-down) line's text."""
        return (self.discipline, self.topic, self.room, self.person)[index]


def _fmt_time(start: str, hours: int, acad_min: int) -> str:
    """Return a ``"08.30-10.00"`` span for a slot starting at ``start``."""
    try:
        begin = datetime.strptime(start, "%H:%M")
    except ValueError:
        return start
    end = begin + timedelta(minutes=hours * acad_min)
    return f"{begin:%H.%M}-{end:%H.%M}"


def _slot_labels(settings: SemesterSettings) -> list[tuple[str, str]]:
    """Return ``(pair-number, time-span)`` captions for every bell slot.

    Pair numbers count academic hours cumulatively, so a two-hour slot reads
    ``"1-2"`` and a following one-hour slot reads ``"5"``.
    """
    labels: list[tuple[str, str]] = []
    hour = 1
    for slot in settings.slots:
        span = slot.hours
        number = str(hour) if span <= 1 else f"{hour}-{hour + span - 1}"
        labels.append((number, _fmt_time(slot.start, span, settings.acad_min)))
        hour += span
    return labels


def _active_days(settings: SemesterSettings) -> list[int]:
    """Return the 0-based indices (Monday=0) of teaching days."""
    return [i for i, on in enumerate(settings.active_days) if on]


def _entities(
    view: str,
    groups: Sequence[Group],
    teachers: Sequence[Teacher],
    majors: Mapping[int, Major],
    entity: str | None,
) -> list[_Entity]:
    """Build the ordered block list for ``view``, optionally scoped to ``entity``.

    ``entity`` scopes the export to a single block: a group *name* in group view
    or a teacher *id* in teacher view (mirroring the client's entity keys).
    """
    if view == "teacher":
        chosen = teachers
        if entity is not None:
            chosen = [t for t in teachers if str(t.id) == str(entity)]
        return [_Entity(key=t.id, caption=t.name) for t in chosen]

    chosen_groups = sorted(groups, key=lambda g: (g.course, g.name))
    if entity is not None:
        chosen_groups = [g for g in chosen_groups if g.name == entity]
    result: list[_Entity] = []
    for group in chosen_groups:
        major = majors.get(group.major_id)
        code = f"\n\n{major.code}" if major else ""
        caption = f"Группа {group.name}{code}\n\nОтветственный\nпреподаватель"
        result.append(_Entity(key=group.id, caption=caption))
    return result


def _cell_text(
    lesson: Lesson,
    view: str,
    disciplines: Mapping[int, Discipline],
    topic_types: Mapping[str, TopicType],
    teachers: Mapping[int, Teacher],
    groups: Mapping[int, Group],
) -> _CellText:
    """Compose the four display lines for a placed ``lesson``."""
    discipline = disciplines.get(lesson.discipline_id) if lesson.discipline_id else None
    kind = topic_types.get(lesson.kind)
    topic = lesson.topic_label.strip()
    if kind is not None:
        topic = f"{topic} ({kind.short})" if topic else kind.label
    if view == "teacher":
        group = groups.get(lesson.group_id)
        person = group.name if group else ""
    else:
        teacher = teachers.get(lesson.sub_by or lesson.teacher_id or 0)
        person = teacher.name if teacher else ""
    return _CellText(
        discipline=discipline.name if discipline else "",
        topic=topic,
        room=lesson.room_id or "",
        person=person,
    )


def _index_lessons(lessons: Sequence[Lesson], view: str) -> dict[tuple[int, int, int, int], Lesson]:
    """Map ``(entity_key, day, slot, week)`` to the placed lesson sitting there.

    The entity key is the group id (group view) or the effective teacher id —
    substitute if present, else staff (teacher view).
    """
    index: dict[tuple[int, int, int, int], Lesson] = {}
    for lesson in lessons:
        if lesson.day is None or lesson.slot is None or lesson.week is None:
            continue
        key = (lesson.sub_by or lesson.teacher_id) if view == "teacher" else lesson.group_id
        if key is None:
            continue
        index[(key, lesson.day, lesson.slot, lesson.week)] = lesson
    return index


def _border(*, left: Side, right: Side, top: Side, bottom: Side) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


def _write(
    ws: Worksheet,
    row: int,
    col: int,
    text: object,
    font: Font,
    border: Border,
    *,
    fill: PatternFill | None = None,
) -> None:
    """Write and fully style a single cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font
    cell.alignment = _CENTER
    cell.border = border
    if fill is not None:
        cell.fill = fill


def _draw_header_row(ws: Worksheet, row: int, view: str, week_dates: Sequence[str]) -> None:
    """Draw a block's caption row: column titles and per-week date headers."""
    last_col = _FIRST_WEEK_COL + len(week_dates) - 1
    _write(
        ws, row, _COL_GROUP,
        "Преподаватель" if view == "teacher" else "Учебная группа",
        _FONT_HEAD, _border(left=_MED, right=_MED, top=_MED, bottom=_MED),
    )
    ws.merge_cells(start_row=row, start_column=_COL_HOUR, end_row=row, end_column=_COL_TIME)
    _write(ws, row, _COL_HOUR, "Учебный час, время", _FONT_HEAD,
           _border(left=_MED, right=_MED, top=_MED, bottom=_MED))
    _write(ws, row, _COL_TIME, None, _FONT_HEAD,
           _border(left=_MED, right=_MED, top=_MED, bottom=_MED))
    for offset, date in enumerate(week_dates):
        col = _FIRST_WEEK_COL + offset
        _write(ws, row, col, date, _FONT_HEAD,
               _border(left=_MED, right=_MED if col == last_col else _THIN, top=_MED, bottom=_MED))


def _draw_block(
    ws: Worksheet,
    top_row: int,
    entity: _Entity,
    day_idx: int,
    view: str,
    slot_labels: Sequence[tuple[str, str]],
    weeks: Sequence[int],
    week_dates: Sequence[str],
    cell_index: Mapping[tuple[int, int, int, int], Lesson],
    cell_text: "_CellRenderer",
) -> int:
    """Render one entity block and return the row after it (incl. spacer)."""
    _draw_header_row(ws, top_row, view, week_dates)

    body_top = top_row + 1
    slot_count = len(slot_labels)
    body_rows = slot_count * _LINES_PER_SLOT
    last_col = _FIRST_WEEK_COL + len(weeks) - 1

    # Column A: one tall merged caption spanning the whole block body.
    ws.merge_cells(start_row=body_top, start_column=_COL_GROUP,
                   end_row=body_top + body_rows - 1, end_column=_COL_GROUP)
    _write(ws, body_top, _COL_GROUP, entity.caption, _FONT_LABEL,
           _border(left=_MED, right=_MED, top=_MED, bottom=_MED))

    for si, (number, time_span) in enumerate(slot_labels):
        slot_top = body_top + si * _LINES_PER_SLOT
        slot_bottom = slot_top + _LINES_PER_SLOT - 1
        first_slot, last_slot = si == 0, si == slot_count - 1
        edge_top = _MED if first_slot else _THIN
        edge_bottom = _MED if last_slot else _THIN

        # Columns B/C: pair number and time, merged over the slot's four lines.
        for col, text in ((_COL_HOUR, number), (_COL_TIME, time_span)):
            ws.merge_cells(start_row=slot_top, start_column=col, end_row=slot_bottom, end_column=col)
            _write(ws, slot_top, col, text, _FONT_HOUR,
                   _border(left=_MED, right=_MED, top=edge_top, bottom=edge_bottom))

        # Week columns: four stacked lines per slot, filled from the schedule.
        for wi, week in enumerate(weeks):
            col = _FIRST_WEEK_COL + wi
            right = _MED if col == last_col else _THIN
            lesson = cell_index.get((entity.key, day_idx, si, week))
            text = cell_text(lesson) if lesson is not None else None
            for line in range(_LINES_PER_SLOT):
                row = slot_top + line
                top = edge_top if line == 0 else _NONE
                bottom = edge_bottom if line == _LINES_PER_SLOT - 1 else _NONE
                _write(
                    ws, row, col,
                    text.line(line) if text is not None else None,
                    _FONT_DISC if line == 0 else _FONT_CELL,
                    _border(left=_MED if col == _FIRST_WEEK_COL else _THIN,
                            right=right, top=top, bottom=bottom),
                    fill=_DATA_FILL,
                )

    for r in range(body_top, body_top + body_rows):
        ws.row_dimensions[r].height = 20.0
    ws.row_dimensions[top_row].height = 30.0
    return body_top + body_rows + 1  # trailing spacer row


class _CellRenderer:
    """Callable that turns a lesson into its :class:`_CellText`, view-aware."""

    def __init__(
        self,
        view: str,
        disciplines: Mapping[int, Discipline],
        topic_types: Mapping[str, TopicType],
        teachers: Mapping[int, Teacher],
        groups: Mapping[int, Group],
    ) -> None:
        self._args = (view, disciplines, topic_types, teachers, groups)

    def __call__(self, lesson: Lesson) -> _CellText:
        view, disciplines, topic_types, teachers, groups = self._args
        return _cell_text(lesson, view, disciplines, topic_types, teachers, groups)


def _title(period: str, year_label: str, majors: Sequence[Major]) -> str:
    """Build the three-line report title shared by every day sheet."""
    codes = ", ".join(m.code for m in majors)
    season = _SEASON_LABEL.get(period, "осеннем")
    vus = f"\nс гражданами, проходящими подготовку по ВУС {codes}" if codes else ""
    return f"РАСПИСАНИЕ ЗАНЯТИЙ{vus}\nв {season} семестре {year_label} учебного года"


def _draw_sheet(
    ws: Worksheet,
    *,
    title: str,
    day_idx: int,
    view: str,
    entities: Sequence[_Entity],
    slot_labels: Sequence[tuple[str, str]],
    weeks: Sequence[int],
    week_dates: Sequence[str],
    cell_index: Mapping[tuple[int, int, int, int], Lesson],
    cell_text: _CellRenderer,
) -> None:
    """Lay out a single day's sheet: title, day name and the entity blocks."""
    last_col = _FIRST_WEEK_COL + len(weeks) - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _write(ws, 1, 1, title, _FONT_TITLE, _border(left=_NONE, right=_NONE, top=_NONE, bottom=_NONE))
    ws.row_dimensions[1].height = 92.0

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    _write(ws, 2, 1, _DAY_NAMES[day_idx], _FONT_DAY,
           _border(left=_NONE, right=_NONE, top=_NONE, bottom=_NONE))
    ws.row_dimensions[2].height = 42.0

    ws.column_dimensions[get_column_letter(_COL_GROUP)].width = 30
    ws.column_dimensions[get_column_letter(_COL_HOUR)].width = 10
    ws.column_dimensions[get_column_letter(_COL_TIME)].width = 16
    for col in range(_FIRST_WEEK_COL, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    row = 4  # rows 1-2 title/day, row 3 breather
    for entity in entities:
        row = _draw_block(
            ws, row, entity, day_idx, view, slot_labels, weeks, week_dates[day_idx], cell_index, cell_text
        )


def build_schedule_workbook(
    *,
    year_label: str,
    period: str,
    view: str,
    settings: SemesterSettings,
    lessons: Sequence[Lesson],
    groups: Sequence[Group],
    teachers: Sequence[Teacher],
    majors: Sequence[Major],
    disciplines: Sequence[Discipline],
    topic_types: Sequence[TopicType],
    entity: str | None = None,
) -> bytes:
    """Build the filled schedule workbook and return it as ``.xlsx`` bytes.

    Args:
        year_label: Academic-year span for the title, e.g. ``"2026-2027"``.
        period: Season key (``"fall"`` / ``"spring"``).
        view: ``"group"`` (block per group) or ``"teacher"`` (block per teacher).
        settings: The season's grid (bell slots, active days, weeks, start date).
        lessons: Lessons of the season; only placed ones are rendered.
        groups: Groups of the year.
        teachers: All teachers.
        majors: Majors, for the title's ВУС codes and group captions.
        disciplines: Disciplines with topics, for the discipline line.
        topic_types: Lesson-type catalogue, for the type short code.
        entity: Optional single-block scope — a group name (group view) or a
            teacher id (teacher view). ``None`` exports every block.

    Returns:
        The serialized ``.xlsx`` document (one sheet per teaching day).
    """
    major_by_id = {m.id: m for m in majors}
    discipline_by_id = {d.id: d for d in disciplines}
    topic_type_by_key = {t.k: t for t in topic_types}
    teacher_by_id = {t.id: t for t in teachers}
    group_by_id = {g.id: g for g in groups}

    entities = _entities(view, groups, teachers, major_by_id, entity)
    slot_labels = _slot_labels(settings)
    day_idxs = _active_days(settings)
    weeks = list(range(1, settings.weeks_count + 1))

    base = _parse_start(settings.start_date)
    week_dates_by_day: dict[int, list[str]] = {
        d: [f"{base + timedelta(days=(w - 1) * 7 + d):%d.%m}" for w in weeks] for d in day_idxs
    }

    cell_index = _index_lessons(lessons, view)
    renderer = _CellRenderer(view, discipline_by_id, topic_type_by_key, teacher_by_id, group_by_id)
    title = _title(period, year_label, majors)

    wb = Workbook()
    wb.remove(wb.active)
    for day_idx in day_idxs:
        ws = wb.create_sheet(title=_DAY_NAMES[day_idx].capitalize())
        _draw_sheet(
            ws,
            title=title,
            day_idx=day_idx,
            view=view,
            entities=entities,
            slot_labels=slot_labels,
            weeks=weeks,
            week_dates=week_dates_by_day,
            cell_index=cell_index,
            cell_text=renderer,
        )
    if not wb.sheetnames:  # no active days configured — keep a valid workbook
        wb.create_sheet(title="Расписание")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_start(start_date: str) -> datetime:
    """Parse the ISO ``yyyy-mm-dd`` semester start; fall back to today."""
    try:
        return datetime.strptime(start_date, "%Y-%m-%d")
    except (ValueError, TypeError):
        return datetime.today()
